import duckdb
from openpyxl import load_workbook
from time import time

BRONZE_PATH = "data/bronze/tabela5938.xlsx"
SILVER_DDL_PATH = "sql/ddl/create_silver.sql"

def create_silver():
    
    START_TIME = time()
    
    conn = duckdb.connect(":memory:")
    conn.execute("INSTALL excel; LOAD excel;")
    
    ##### Bronze (Temp) #####
    conn.execute("CREATE SCHEMA IF NOT EXISTS bronze")
    
    wb = load_workbook(filename=BRONZE_PATH, read_only=True, keep_links=False)
    sheets = {
        "pib": wb.sheetnames[0],
        "imp": wb.sheetnames[1],
        "valor_add": wb.sheetnames[2]
    }
    
    for k, v in sheets.items():
        bronze_table = f"bronze.{k}"
        bronze_sheet = v
        
        conn.execute(f"""
            CREATE OR REPLACE TABLE {bronze_table} AS
            SELECT 
                "Unidade da Federação e Município" AS localidade,
                Ano AS ano_2002,
                C2 AS ano_2003,
                "_1" AS ano_2004, "_2" AS ano_2005, "_3" AS ano_2006, "_4" AS ano_2007, "_5" AS ano_2008,
                "_6" AS ano_2009, "_7" AS ano_2010, "_8" AS ano_2011, "_9" AS ano_2012, "_10" AS ano_2013,
                "_11" AS ano_2014, "_12" AS ano_2015, "_13" AS ano_2016, "_14" AS ano_2017, "_15" AS ano_2018,
                "_16" AS ano_2019, "_17" AS ano_2020, "_18" AS ano_2021, "_19" AS ano_2022, "_20" AS ano_2023
            FROM read_xlsx(
                '{BRONZE_PATH}',
                sheet='{bronze_sheet}',
                range='A3:W5601',
                header=true,
                all_varchar=true
            );
        """)
    
    ##### Silver #####
    with open(SILVER_DDL_PATH, mode="r") as f:
        conn.execute(f.read())
    
    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS silver.dim_municipio AS (
            WITH municipios AS (
                SELECT
                    SUBSTRING(localidade FROM 1 FOR LENGTH(localidade) - 5) AS nome_municipio,
                    RIGHT(localidade, 3)[1:2] AS sigla_uf
                FROM bronze.pib
                WHERE localidade LIKE '%)'
            )
            
            SELECT
                ROW_NUMBER() OVER(
                    ORDER BY m.nome_municipio, u.uf_id
                ) AS municipio_id,
                m.nome_municipio,
                u.uf_id
            FROM municipios m
            JOIN silver.dim_uf u
                ON u.sigla_uf = m.sigla_uf
            GROUP BY m.nome_municipio, u.uf_id
        )
    """)
    
    anos_cols = [f"ano_{y}" for y in range(2002, 2024)]
    
    for s in sheets.keys():
        bronze_table = f"bronze.{s}"
        silver_table = f"silver.fact_{s}"
        valor_col = f"valor_{s}"
        
        conn.execute(f"""
            CREATE OR REPLACE TABLE {silver_table} AS (
                WITH bronze_unpivot AS (
                    SELECT *
                    FROM {bronze_table}
                    UNPIVOT (valor FOR coluna_ano IN ({", ".join(anos_cols)}))
                )
                
                SELECT
                    m.municipio_id,
                    t.ano_id,
                    CASE
                        WHEN valor = '...' THEN NULL
                        ELSE CAST(valor AS DOUBLE)
                    END AS {valor_col}
                FROM bronze_unpivot u
                JOIN silver.dim_municipio m
                    ON m.nome_municipio =
                       SUBSTRING(u.localidade FROM 1 FOR LENGTH(u.localidade) - 5)
                JOIN silver.dim_tempo t
                    ON t.ano = CAST(REPLACE(u.coluna_ano, 'ano_', '') AS INTEGER)
                WHERE u.localidade LIKE '%)'
            )
        """)
        
        conn.execute(f"DROP TABLE {bronze_table}")
    
    conn.execute("DROP SCHEMA bronze")
    
    END_TIME = time()
    
    print(f"[SUCCESS] Bronze -> Silver process finished successfully in {END_TIME - START_TIME:.2f} seconds!")